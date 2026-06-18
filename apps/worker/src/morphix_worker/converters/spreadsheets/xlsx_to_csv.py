from __future__ import annotations

import csv
from pathlib import Path

from ..common import expected_output


class XlsxToCsvConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        if target_format != "csv":
            raise ValueError("XLSX converter only supports CSV output")

        from openpyxl import load_workbook

        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheet = workbook.active
        output_path = expected_output(input_path, output_dir, target_format)
        with output_path.open("w", newline="", encoding="utf-8") as output_file:
            writer = csv.writer(output_file)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        workbook.close()
        return output_path

