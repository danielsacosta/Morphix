from __future__ import annotations

import csv
import shutil
import subprocess
from pathlib import Path
from typing import Protocol


SUPPORTED_PAIRS: set[tuple[str, str]] = {
    ("docx", "pdf"),
    ("xlsx", "pdf"),
    ("pptx", "pdf"),
    ("pdf", "docx"),
    ("csv", "xlsx"),
    ("xlsx", "csv"),
    ("png", "jpg"),
    ("jpg", "png"),
    ("jpg", "webp"),
    ("png", "webp"),
    ("mp4", "mp3"),
    ("mp4", "webm"),
    ("mov", "mp4"),
    ("wav", "mp3"),
    ("mp3", "wav"),
}


class Converter(Protocol):
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        ...


def run_command(command: list[str], timeout_seconds: int) -> None:
    subprocess.run(command, check=True, timeout=timeout_seconds, stdout=subprocess.PIPE, stderr=subprocess.PIPE)


def expected_output(input_path: Path, output_dir: Path, target_format: str) -> Path:
    extension = "jpg" if target_format == "jpeg" else target_format
    return output_dir / f"{input_path.stem}.{extension}"


class LibreOfficePdfConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        if target_format != "pdf":
            raise ValueError("LibreOffice converter only supports PDF output")

        run_command(
            [
                "libreoffice",
                "--headless",
                "--nologo",
                "--nofirststartwizard",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(input_path),
            ],
            timeout_seconds,
        )
        output_path = expected_output(input_path, output_dir, target_format)
        if not output_path.exists():
            raise FileNotFoundError(f"LibreOffice did not create {output_path}")
        return output_path


class PdfToDocxConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        if target_format != "docx":
            raise ValueError("PDF converter only supports DOCX output")

        from pdf2docx import Converter as PdfConverter

        output_path = expected_output(input_path, output_dir, target_format)
        converter = PdfConverter(str(input_path))
        try:
            converter.convert(str(output_path), start=0, end=None)
        finally:
            converter.close()
        if not output_path.exists():
            raise FileNotFoundError(f"pdf2docx did not create {output_path}")
        return output_path


class CsvToXlsxConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        if target_format != "xlsx":
            raise ValueError("CSV converter only supports XLSX output")

        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        with input_path.open(newline="", encoding="utf-8-sig") as source_file:
            reader = csv.reader(source_file)
            for row in reader:
                worksheet.append(row)

        output_path = expected_output(input_path, output_dir, target_format)
        workbook.save(output_path)
        return output_path


class XlsxToCsvConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        if target_format != "csv":
            raise ValueError("XLSX converter only supports CSV output")

        from openpyxl import load_workbook

        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheet = workbook.active
        output_path = expected_output(input_path, output_dir, target_format)
        with output_path.open("w", newline="", encoding="utf-8") as output_file:
            writer = csv.writer(output_file)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
        workbook.close()
        return output_path


class ImageMagickConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        binary = shutil.which("magick") or shutil.which("convert")
        if not binary:
            raise FileNotFoundError("ImageMagick binary not found")

        output_path = expected_output(input_path, output_dir, target_format)
        command = [binary, str(input_path), str(output_path)]
        run_command(command, timeout_seconds)
        if not output_path.exists():
            raise FileNotFoundError(f"ImageMagick did not create {output_path}")
        return output_path


class FFmpegConverter:
    def convert(self, input_path: Path, output_dir: Path, target_format: str, timeout_seconds: int) -> Path:
        output_path = expected_output(input_path, output_dir, target_format)
        run_command(["ffmpeg", "-y", "-i", str(input_path), str(output_path)], timeout_seconds)
        if not output_path.exists():
            raise FileNotFoundError(f"FFmpeg did not create {output_path}")
        return output_path


class ConverterRegistry:
    def __init__(self) -> None:
        libreoffice = LibreOfficePdfConverter()
        imagemagick = ImageMagickConverter()
        ffmpeg = FFmpegConverter()
        self._converters: dict[tuple[str, str], Converter] = {
            ("docx", "pdf"): libreoffice,
            ("xlsx", "pdf"): libreoffice,
            ("pptx", "pdf"): libreoffice,
            ("pdf", "docx"): PdfToDocxConverter(),
            ("csv", "xlsx"): CsvToXlsxConverter(),
            ("xlsx", "csv"): XlsxToCsvConverter(),
            ("png", "jpg"): imagemagick,
            ("jpg", "png"): imagemagick,
            ("jpg", "webp"): imagemagick,
            ("png", "webp"): imagemagick,
            ("mp4", "mp3"): ffmpeg,
            ("mp4", "webm"): ffmpeg,
            ("mov", "mp4"): ffmpeg,
            ("wav", "mp3"): ffmpeg,
            ("mp3", "wav"): ffmpeg,
        }

    @property
    def supported_pairs(self) -> set[tuple[str, str]]:
        return set(self._converters)

    def convert(self, source_format: str, target_format: str, input_path: Path, output_dir: Path, timeout_seconds: int) -> Path:
        key = (source_format, target_format)
        converter = self._converters.get(key)
        if not converter:
            raise ValueError(f"Unsupported conversion {source_format} to {target_format}")
        return converter.convert(input_path, output_dir, target_format, timeout_seconds)

